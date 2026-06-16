"""Importa repostajes desde el Excel de seguimiento de combustible.

Cada hoja cuyo nombre parece una matrícula se importa como un vehículo.
Columnas esperadas (fila 0 cabecera): Fecha, Estación, Importe, Volumen, Odómetro.

Uso (en VestraApp, con .env cargado):
    /opt/vestra/venv/bin/python scripts/import_fuel_excel.py \
        --file /tmp/combustible.xlsx --user susoinc@gmail.com

Idempotente: no duplica repostajes ya existentes (clave: vehículo+fecha+importe+odómetro).
"""
from __future__ import annotations

import argparse
import sys
import uuid
from datetime import date
from decimal import Decimal

import openpyxl

sys.path.insert(0, __import__("os").path.dirname(__import__("os").path.dirname(
    __import__("os").path.abspath(__file__))))

from app import create_app  # noqa: E402
from app.extensions import db  # noqa: E402
from app.models.user import User  # noqa: E402
from app.models.vehicle import Vehicle, FuelLog  # noqa: E402
from sqlalchemy import select  # noqa: E402


def looks_like_plate(name: str) -> bool:
    s = name.replace(" ", "")
    return 5 <= len(s) <= 9 and any(c.isdigit() for c in s) and any(c.isalpha() for c in s)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True)
    ap.add_argument("--user", required=True)
    args = ap.parse_args()

    app = create_app()
    with app.app_context():
        user = db.session.execute(select(User).where(User.email == args.user)).scalars().first()
        if not user:
            print(f"Usuario {args.user} no encontrado", file=sys.stderr)
            sys.exit(1)

        wb = openpyxl.load_workbook(args.file, data_only=True)
        total_new = 0
        for sheet in wb.sheetnames:
            if not looks_like_plate(sheet):
                print(f"· hoja '{sheet}' ignorada (no parece matrícula)")
                continue
            plate = sheet.strip().upper()

            vehicle = db.session.execute(
                select(Vehicle).where(Vehicle.user_id == user.id, Vehicle.plate == plate)
            ).scalars().first()
            if not vehicle:
                vehicle = Vehicle(id=str(uuid.uuid4()), user_id=user.id,
                                  nickname=plate, plate=plate, status="active")
                db.session.add(vehicle)
                db.session.flush()
                print(f"+ vehículo creado: {plate}")

            ws = wb[sheet]
            new_here = 0
            max_odo = vehicle.current_km or 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue  # cabecera
                d, station, importe, volumen, odo = (list(row) + [None] * 5)[:5]
                if not d or importe is None:
                    continue
                log_date = d.date() if hasattr(d, "date") else d
                if not isinstance(log_date, date):
                    continue
                cost = Decimal(str(round(float(importe), 2)))
                liters = Decimal(str(round(float(volumen), 2))) if volumen is not None else None
                odo_km = int(odo) if odo is not None else None

                exists = db.session.execute(
                    select(FuelLog).where(
                        FuelLog.vehicle_id == vehicle.id,
                        FuelLog.log_date == log_date,
                        FuelLog.total_cost == cost,
                    )
                ).scalars().first()
                if exists:
                    continue

                ppl = round(cost / liters, 4) if liters else None
                db.session.add(FuelLog(
                    id=str(uuid.uuid4()), vehicle_id=vehicle.id, log_date=log_date,
                    station=str(station).strip() if station else None,
                    liters=liters, total_cost=cost, odometer_km=odo_km,
                    price_per_liter=ppl, full_tank=True,
                ))
                new_here += 1
                if odo_km and odo_km > max_odo:
                    max_odo = odo_km

            vehicle.current_km = max_odo or vehicle.current_km
            total_new += new_here
            print(f"  {plate}: {new_here} repostajes nuevos")

        db.session.commit()
        print(f"\nListo. {total_new} repostajes importados.")


if __name__ == "__main__":
    main()
